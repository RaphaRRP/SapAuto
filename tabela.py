import openpyxl as xl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import NamedStyle
from datetime import datetime
import pyperclip
import pandas as pd
import estilo
import re


data_hoje = datetime.now().strftime("%d-%m")


arquivo_mb25 = f"C:/Users/mis7ca/Desktop/Projetos/SapAuto/tabelas/mb25/mb25 - {data_hoje}.xlsx"
arquivo_me3m = f"C:/Users/mis7ca/Desktop/Projetos/SapAuto/tabelas/me3m/me3m - {data_hoje}.xlsx"
arquivo_z22k032 = f"C:/Users/mis7ca/Desktop/Projetos/SapAuto/tabelas/z22k032/z22k032 - {data_hoje}.xlsx"
arquivo_z22m051 = f"C:/Users/mis7ca/Desktop/Projetos/SapAuto/tabelas/z22m051/z22m051 - {data_hoje}.xlsx"
arquivo_blocok_depois = "C:/Users/mis7ca/Desktop/Projetos/SapAuto/tabelas/Bloco K.xlsx"



def excluir_linhas(tabela, coluna, valores): #Recebe Coluna e Valor, exclui as linhas que o valor esta na coluna
    linhas_a_remover = []
    for row in range(2, tabela.max_row + 1):  # Começa da linha 2 para evitar o cabeçalho
        codigo_valor = tabela[f'{coluna}{row}'].value
        if codigo_valor in valores:
            linhas_a_remover.append(row)
    for row in reversed(linhas_a_remover):
        tabela.delete_rows(row, 1)


def transformar_coluna_em_texto(tabela, coluna):
    for row in range(2, tabela.max_row + 1):  
        cell = tabela[f'{coluna}{row}']
        cell.number_format = '@'

def transformar_texto_051(tabela, coluna):
    for row in range(2, tabela.max_row + 1):  
        cell = tabela[f'{coluna}{row}']
        cell.value = re.sub(r"7188006", "7188.006.", cell.value)
        cell.value = re.sub(r"7188007", "7188.007.", cell.value)


def transformar_coluna_em_data_abreviada(tabela, coluna):
    for row in range(2, tabela.max_row + 1):
        cell = tabela[f'{coluna}{row}']
        # Altera o formato da célula para data abreviada
        cell.number_format = 'DD/MM/YY'


def alterar_valor(tabela, coluna, valor_antes, valor_depois):
    for row in range(2, tabela.max_row + 1):  
        cell = tabela[f'{coluna}{row}']
        if cell.value == valor_antes:
            cell.value = valor_depois


#========================= Bloco K =======================#

def copiar_tabela_entre_arquivos(arquivo_origem, nome_planilha_origem, arquivo_destino, nome_planilha_destino=None):
    df = pd.read_excel(arquivo_origem, sheet_name=nome_planilha_origem)
    with pd.ExcelWriter(arquivo_destino, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=nome_planilha_destino, index=False)


def copiar_lista_material():

    df_origem = pd.read_excel(arquivo_blocok_depois, sheet_name='MB25')
    df_destino = pd.read_excel(arquivo_blocok_depois, sheet_name='Planilha1')

    coluna_origem = df_origem[['Material']]
    df_destino = coluna_origem.drop_duplicates(subset=['Material'])

    with pd.ExcelWriter(arquivo_blocok_depois, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_origem.to_excel(writer, sheet_name='MB25', index=False)
        df_destino.to_excel(writer, sheet_name='Planilha1', index=False)



def pegar_partnumbers():
    wb = xl.load_workbook(arquivo_blocok_depois)
    ws = wb['Planilha1']
    parnumbers = []

    for cell in ws['A'][1:]:
        if cell.value is not None:
            parnumbers.append(cell.value)
    
    parnumbers_str = "\r\n".join(parnumbers)
    pyperclip.copy(parnumbers_str)

    return parnumbers_str


#========================= MB25 =======================#

def levar_mb25_para_blocok():

    wb_mb25 = xl.load_workbook(arquivo_mb25)
    tabela_mb25 = wb_mb25['Sheet1']
    header_mb25 = {cell.value: cell.column for cell in tabela_mb25[1]}

    deposito_coluna_idx_mb25 = header_mb25['Depósito']
    deposito_coluna_mb25 = xl.utils.get_column_letter(deposito_coluna_idx_mb25)
    transformar_coluna_em_texto(tabela_mb25, deposito_coluna_mb25)
    excluir_linhas(tabela_mb25, deposito_coluna_mb25, ['9010', '9032', ''])

    UM_coluna_idx_mb25 = header_mb25['UM básica']
    UM_coluna_mb25 = xl.utils.get_column_letter(UM_coluna_idx_mb25)
    excluir_linhas(tabela_mb25, UM_coluna_mb25, ['KG'])

    wb_mb25.save(arquivo_mb25)
    
    copiar_tabela_entre_arquivos(arquivo_mb25, 'Sheet1', arquivo_blocok_depois,  'MB25')
    copiar_lista_material()
    pegar_partnumbers()
    

    
def alterar_tabelas():

    copiar_tabela_entre_arquivos(arquivo_me3m, 'Sheet1', arquivo_blocok_depois,  'ME3M Programa Remessa')
    copiar_tabela_entre_arquivos(arquivo_z22k032, 'Sheet1', arquivo_blocok_depois,  'saldo z22k032')
    copiar_tabela_entre_arquivos(arquivo_z22m051, 'Sheet1', arquivo_blocok_depois,  'Z22M051 Programação')

    estilo.formatar_blocok(arquivo_blocok_depois)


    #========================== mb25 ==================================#
    wb_mb25 = xl.load_workbook(arquivo_blocok_depois)
    tabela_mb25 = wb_mb25['MB25']
    header_mb25 = {cell.value: cell.column for cell in tabela_mb25[1]}
    data_coluna_idx_mb25 = header_mb25['Data da necessidade']
    data_coluna_mb25 = xl.utils.get_column_letter(data_coluna_idx_mb25)
    transformar_coluna_em_data_abreviada(tabela_mb25, data_coluna_mb25)

    material_coluna_idx_mb25 = header_mb25['Material']
    material_coluna_mb25 = xl.utils.get_column_letter(material_coluna_idx_mb25)
    transformar_coluna_em_texto(tabela_mb25, material_coluna_mb25)

    data_coluna_idx_mb25 = header_mb25['Data da necessidade']
    data_coluna_mb25 = xl.utils.get_column_letter(data_coluna_idx_mb25)
    transformar_coluna_em_data_abreviada(tabela_mb25, data_coluna_mb25)

    wb_mb25.save(arquivo_blocok_depois)

    #========================== me3m ==================================#
    wb_me3m = xl.load_workbook(arquivo_blocok_depois)
    tabela_me3m = wb_me3m['ME3M Programa Remessa']
    header_me3m = {cell.value: cell.column for cell in tabela_me3m[1]}

    codigo_coluna_idx_me3m = header_me3m['Código de eliminação']
    codigo_coluna_me3m = xl.utils.get_column_letter(codigo_coluna_idx_me3m)
    excluir_linhas(tabela_me3m, codigo_coluna_me3m, ['S'])

    materiais_coluna_idx_me3m = header_me3m['Material']
    materiais_coluna_me3m = xl.utils.get_column_letter(materiais_coluna_idx_me3m)
    transformar_coluna_em_texto(tabela_me3m, materiais_coluna_me3m)

    data_coluna_idx_me3m = header_me3m['Data do documento']
    data_coluna_me3m = xl.utils.get_column_letter(data_coluna_idx_me3m)
    transformar_coluna_em_data_abreviada(tabela_me3m, data_coluna_me3m)

    wb_me3m.save(arquivo_blocok_depois)


    #========================== z22k032 ==================================#
    wb_z22k032 = xl.load_workbook(arquivo_blocok_depois)
    tabela_z22k032 = wb_z22k032['saldo z22k032']
    header_z22k032 = {cell.value: cell.column for cell in tabela_z22k032[1]}

    materiais_coluna_idx_z22k032 = header_z22k032['Material']
    materiais_coluna_z22k032 = xl.utils.get_column_letter(materiais_coluna_idx_z22k032)
    transformar_coluna_em_texto(tabela_z22k032, materiais_coluna_z22k032)
    
    wb_z22k032.save(arquivo_blocok_depois)

    #========================== z22m051 ==================================#
    wb_z22m051 = xl.load_workbook(arquivo_blocok_depois)
    tabela_z22m051 = wb_z22m051['Z22M051 Programação']
    header_z22m051 = {cell.value: cell.column for cell in tabela_z22m051[1]}

    UM_coluna_idx_z22m051 = header_z22m051['UM']
    UM_coluna_z22m051 = xl.utils.get_column_letter(UM_coluna_idx_z22m051)
    alterar_valor(tabela_z22m051, UM_coluna_z22m051, 'ST', 'PC')

    materiais_coluna_idx_z22m051 = header_z22m051['Material']
    materiais_coluna_z22m051 = xl.utils.get_column_letter(materiais_coluna_idx_z22m051)
    transformar_texto_051(tabela_z22m051, materiais_coluna_z22m051)
    transformar_coluna_em_texto(tabela_z22m051, materiais_coluna_z22m051)

    data_coluna_idx_z22m051 = header_z22m051['Data Remessa']
    data_coluna_z22m051 = xl.utils.get_column_letter(data_coluna_idx_z22m051)
    transformar_coluna_em_data_abreviada(tabela_z22m051, data_coluna_z22m051)

    wb_z22m051.save(arquivo_blocok_depois)

    #========================== Bloco K ==================================#

