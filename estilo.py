import openpyxl as xl
from openpyxl.styles import PatternFill, Border, Side, NamedStyle, Font
from openpyxl.utils.dataframe import dataframe_to_rows

#Para as tabelas de mb25
def formatar_blocok(arquivo):

    wb = xl.load_workbook(arquivo)

    tabela_mb25 = wb["MB25"]
    tabela_planilha1 = wb["Planilha1"]
    tabelas = [tabela_mb25, tabela_planilha1]

    for tabela in tabelas:

        ajustar_largura_colunas(tabela)
        estilizar_cabecalho(tabela)
        wb.save(arquivo)


def ajustar_largura_colunas(planilha):

    for coluna in planilha.columns:
        max_length = 0
        letra_coluna = coluna[0].column_letter  

        for celula in coluna:
            try:
                if len(str(celula.value)) > max_length:
                    max_length = len(str(celula.value))
            except:
                pass
        
        largura_ajustada = (max_length + 2)  
        planilha.column_dimensions[letra_coluna].width = largura_ajustada


def estilizar_cabecalho(planilha):
    fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    border = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )
    font = Font(bold=True)
    for celula in planilha[1]:
        celula.fill = fill
        celula.border = border
        celula.font = font

